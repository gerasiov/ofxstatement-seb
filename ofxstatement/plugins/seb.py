# -*- coding: utf-8 -*-

import re
import itertools
import logging

from datetime import datetime
from openpyxl import load_workbook

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id


def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(itertools.islice(iterable, n))


class SebStatementParser(StatementParser):
    date_format = '%Y-%m-%d'
    bank_id = 'SEB'
    currency_id = 'SEK'
    footer_regexps = [
        '^Datum:  -',
        '^Datum: ([0-9]{4}-[0-9]{2}-[0-9]{2}) - ([0-9]{4}-[0-9]{2}-[0-9]{2})$'
    ]

    def __init__(self, fin, brief=False):
        """
        Create a new SebStatementParser instance.

        :param fin: filename to create parser for
        :param brief: whenever to attempt replace description with a brief version i.e. all extra info removed
        """

        self.workbook = load_workbook(filename=fin, read_only=True)
        self.brief = brief

        self.validate()
        self.statement = self.parse_statement()

    def validate(self):
        """
        Naive validation to make sure that xlsx document is structured the way it was
        when this parser was written.

        :raises ValueError if workbook has invalid format
        """

        try:
            self._validate()
        except AssertionError as e:
            raise ValueError(e)

    def _validate(self):
        sheet = self.workbook.active

        logging.info('Checking that sheet has at least 5 rows.')
        rows = take(5, sheet.iter_rows())
        assert len(rows) == 5

        logging.info('Extracting values for every cell.')
        rows = [[c.value for c in row] for row in rows]

        logging.info('Verifying summary header.')
        summary_header_row = rows[0]
        assert ['Saldo', 'Disponibelt belopp', 'Beviljad kredit', None, None] == summary_header_row[1:]

        logging.info('Getting account id.')
        summary_account_row = rows[1]
        account_id = summary_account_row[0]

        def is_footer(row):
            for r in self.footer_regexps:
                if re.match(r, row[0]):
                    return True
            return False

        logging.info('Verifying summary footer.')
        summary_footer_row = rows[2]
        assert is_footer(summary_footer_row)
        assert [None, None, None, None, None] == summary_footer_row[1:]

        logging.info('Skipping empty/padding row.')
        empty_row = rows[3]
        assert [None, None, None, None, None, None] == empty_row

        logging.info('Verifying statements header.')
        statement_header_row = rows[4]
        assert re.match('^Bokföringsdatum$', statement_header_row[0])
        assert re.match('^Valutadatum$', statement_header_row[1])
        assert re.match('^Verifikationsnummer$', statement_header_row[2])
        assert re.match('^Text / mottagare$', statement_header_row[3])
        assert re.match('^Belopp$', statement_header_row[4])
        assert re.match('^Saldo$', statement_header_row[5])

        logging.info('Everything is OK!')

    def parse_statement(self):
        statement = Statement()
        sheet = self.workbook.active

        # We need only first 3 rows here.
        rows = take(3, sheet.iter_rows())
        rows = [[c.value for c in row] for row in rows]

        assert len(rows) == 3
        header_row, account_row, footer_row = rows

        account_id, saldo, disponibelt_belopp, beviljad_kredit, _1, _2 = account_row
        statement.account_id = account_id
        statement.end_balance = float(saldo)
        statement.bank_id = self.bank_id
        statement.currency = self.currency_id

        for r in self.footer_regexps:
            m = re.match(r, footer_row[0])
            if m and m.groups():
                part_from, part_to = m.groups()
                statement.start_date = self.parse_datetime(part_from)
                statement.end_date = self.parse_datetime(part_to)

        return statement

    def split_records(self):
        sheet = self.workbook.active

        # Skip first 5 rows. Headers they are.
        for row in itertools.islice(sheet.iter_rows(), 5, None):
            yield [c.value for c in row]

    def parse_record(self, row):
        row = take(5, row)

        stmt_line = StatementLine()
        stmt_line.date = self.parse_datetime(row[0])
        _ = self.parse_datetime(row[1])  # TODO: ???
        stmt_line.refnum = row[2]
        stmt_line.memo = row[3]
        stmt_line.amount = row[4]

        #
        # Looks like SEB formats description for card transactions so it includes the actual purchase date
        # within e.g. 'WIRSTRÖMS PU/14-12-31' and it means that description is 'WIRSTRÖMS PU' while the actual
        # card operation is 2014-12-31.
        #
        # P.S. Wirströms Irish Pub is our favorite pub in Stockholm.
        #
        m = re.match('(.*)/([0-9]{2}-[0-9]{2}-[0-9]{2})$', stmt_line.memo)
        if m:
            card_memo, card_date = m.groups()
            if self.brief:
                stmt_line.memo = card_memo
            stmt_line.date_user = datetime.strptime(card_date, '%y-%m-%d')

        stmt_line.id = generate_transaction_id(stmt_line)
        return stmt_line


def parse_bool(value):
    if value in ('True', 'true', '1'):
        return True
    if value in ('False', 'false', '0'):
        return False
    raise ValueError("Can't parse boolean value: %s" % value)


class SebPlugin(Plugin):
    def get_parser(self, fin):
        kwargs = {}
        if self.settings:
            if 'brief' in self.settings:
                kwargs['brief'] = parse_bool(self.settings.get('brief'))
        return SebStatementParser(fin, **kwargs)
