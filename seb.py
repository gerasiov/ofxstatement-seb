# -*- coding: utf-8 -*-

import re
import itertools

from datetime import datetime
from openpyxl import load_workbook

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement.statement import Statement, StatementLine


def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(itertools.islice(iterable, n))


def validate_workbook(workbook):
    """
    Naive validation to make sure that xlsx document is structured the way it was
    when this parser was written.

    :raises ValueError if workbook has invalid format
    """

    sheet = workbook.get_active_sheet()
    try:
        # We need only first 5 rows.
        rows = take(5, sheet.iter_rows())

        header = [c.value for c in rows[0]]
        assert header[0] == 'Privatkonto'
        assert header[1] == 'Saldo'
        assert header[2] == 'Disponibelt belopp'
        assert header[3] == 'Beviljad kredit'
        assert header[4] is None
        assert header[5] is None

        header = [c.value for c in rows[2]]
        assert header[0].startswith('Datum:')
        assert header[1] is None
        assert header[2] is None
        assert header[3] is None
        assert header[4] is None
        assert header[5] is None

        header = [c.value for c in rows[3]]
        assert header[0].startswith('Bokf')  # Bokförings-
        assert header[1].startswith('Valuta-')
        assert header[2].startswith('Verifikations-')
        assert header[3] is None
        assert header[4] is None
        assert header[5] is None

        header = [c.value for c in rows[4]]
        assert header[0] is None
        assert header[1] is None
        assert header[2] is None
        assert header[3].startswith('Text / mottagare')
        assert header[4].startswith('Belopp')
        assert header[5].startswith('Saldo')
    except AssertionError as e:
        raise ValueError(e)


class SebStatementParser(StatementParser):
    date_format = '%Y-%m-%d'
    bank_id = 'SEB'
    currency_id = 'SEK'

    def __init__(self, fin):
        self.workbook = load_workbook(filename=fin, read_only=True)
        validate_workbook(self.workbook)
        self.statement = self.parse_statement()

    def parse_statement(self):
        statement = Statement()
        sheet = self.workbook.get_active_sheet()

        # We need only first 2 rows here.
        rows = take(3, sheet.iter_rows())

        values = [c.value for c in rows[1]]
        privatkonto, saldo, disponibelt_belopp, beviljad_kredit, _1, _2 = values
        statement.account_id = privatkonto
        statement.end_balance = float(saldo)
        statement.bank_id = self.bank_id
        statement.currency = self.currency_id

        header = rows[2]
        date_regexp = '[0-9]{4}-[0-9]{2}-[0-9]{2}'
        m = re.match('^Datum: (%s) - (%s)$' % (date_regexp, date_regexp), header[0].value)
        if m:
            part_from, part_to = m.groups()
            statement.start_date = self.parse_datetime(part_from)
            statement.end_date = self.parse_datetime(part_to)

        return statement

    def split_records(self):
        sheet = self.workbook.get_active_sheet()

        # Skip first 5 rows. Headers they are.
        for row in itertools.islice(sheet.iter_rows(), 5, None):
            yield (c.value for c in row)

    def parse_record(self, row):
        row = take(5, row)

        stmt_line = StatementLine()
        stmt_line.date = self.parse_datetime(row[0])
        _ = self.parse_datetime(row[1])
        stmt_line.id = row[2]
        stmt_line.memo = row[3]
        stmt_line.amount = row[4]

        #
        # Looks like SEB formats description for card transactions so it includes the actual purchase date
        # within e.g. 'WIRSTRÖMS PU/14-12-31' and it means that description is 'WIRSTRÖMS PU' while the actual
        # card operation is 2014-12-31.
        #
        # P.S. Wirströms Irish Pub is our favorite pub in Stockholm.
        #
        m = re.match('(.*)/([0-9]{2}-[0-9]{2}-[0-9]{2})$', stmt_line.memo)
        if m:
            stmt_line.memo, date_string = m.groups()
            stmt_line.date_user = datetime.strptime(date_string, '%y-%m-%d')

        return stmt_line


class SebPlugin(Plugin):
    def get_parser(self, fin):
        return SebStatementParser(fin)
